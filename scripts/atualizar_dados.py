#!/usr/bin/env python3
"""
Atualiza data/dataset.js a partir da planilha Excel do projeto.

Uso:
    python3 scripts/atualizar_dados.py

Requisitos:
    pip install openpyxl

O que faz:
    1. Abre "Saida_BEL.xlsx" (na raiz do projeto) e lê a aba "jan_26 (2)".
    2. Limpa e converte os lançamentos para o formato usado pelo dashboard.
    3. Regrava data/dataset.js com os dados novos.

Depois de rodar, basta atualizar a página do dashboard no navegador
(F5) para ver os dados novos — não precisa reiniciar nada.
"""

import datetime
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERRO: a biblioteca 'openpyxl' não está instalada.")
    print("Rode primeiro:  pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Configuração — ajuste aqui se o nome do arquivo ou da aba mudar
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = PROJECT_ROOT / "Saida_BEL.xlsx"
SHEET_NAME = "jan_26 (2)"
OUTPUT_PATH = PROJECT_ROOT / "data" / "dataset.js"


def to_seconds(value):
    """Converte horários/durações do Excel para segundos (ou None)."""
    if value is None:
        return None
    if isinstance(value, datetime.time):
        return value.hour * 3600 + value.minute * 60 + value.second
    if isinstance(value, datetime.timedelta):
        return value.total_seconds()
    if isinstance(value, (int, float)):
        return value * 86400
    return None


def normalize_occurrence(raw):
    """Converte o valor bruto da coluna 'Ocorrência' em uma lista de códigos de 2 dígitos.

    A coluna aceita múltiplos códigos separados por vírgula (ex.: "38, 13"). Como a planilha
    usa configuração regional pt-BR, quando alguém digita dois códigos separados por vírgula
    sem espaço o Excel às vezes interpreta a vírgula como separador decimal e guarda um único
    número (ex.: "38,13" vira 38.13). Esta função desfaz essa conversão, recompondo os códigos
    originais a partir da parte inteira e da parte decimal.
    """
    if raw is None:
        return []
    if isinstance(raw, bool):
        return []
    if isinstance(raw, int):
        return [] if raw == 0 else [str(raw).zfill(2)]
    if isinstance(raw, float):
        if raw == 0:
            return []
        texto = f"{raw:.2f}"  # ex.: 42.1 -> "42.10", 9.13 -> "9.13"
        inteiro, decimal = texto.split(".")
        return [inteiro.zfill(2), decimal.zfill(2)]
    if isinstance(raw, str):
        raw = raw.strip()
        if raw in ("", "0"):
            return []
        partes = [p.strip() for p in raw.split(",")]
        return [p.zfill(2) for p in partes if p not in ("", "0")]
    return []


def clean_records(ws):
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    headers = [c.value for c in ws[1]]

    def col(name):
        return headers.index(name)

    records = []
    for r in rows:
        d = dict(zip(headers, r))
        if d.get("Data") is None or d.get("Placa") is None:
            continue  # ignora linhas vazias/incompletas

        rec = {}
        rec["data"] = d["Data"].strftime("%Y-%m-%d")
        rec["ano"] = d["Ano"]
        rec["mes"] = d["Mês"]
        rec["placa"] = d["Placa"]
        rec["motorista"] = d.get("Motorista") or "N/D"
        rec["valorFrete"] = float(d.get("Valor Frete") or 0)
        rec["valorMercadoria"] = float(d.get("Valor Mercadoria") or 0)
        rec["peso"] = float(d.get("Peso") or 0)
        rec["pctFrete"] = float(d.get("% Frete") or 0)
        rec["hrSaida"] = to_seconds(d.get("Hr Saida"))
        rec["hrChegada"] = to_seconds(d.get("Hr Chegada"))
        rec["tempoSeg"] = to_seconds(d.get("Tempo"))
        rec["kmSaida"] = d.get("Km Saída")
        rec["kmChegada"] = d.get("Km Chegada")
        rec["kmDia"] = float(d.get("Km/Dia") or 0)
        rec["vols"] = float(d.get("Vols") or 0)
        rec["entregas"] = float(d.get("Entregas") or 0)
        rec["realizadas"] = float(d.get("Realizadas") or 0)
        rec["retornadas"] = float(d.get(" Retornadas") or 0)
        rec["pctEntrega"] = float(d.get("% Entrega") or 0)
        rec["meta"] = d.get("Meta") or "N/D"
        rec["cidade"] = d.get("Cidade Destino") or "N/D"

        nfs = d.get("Nfs Voltaram")
        rec["nfsVoltaram"] = float(nfs) if isinstance(nfs, (int, float)) else 0

        occ = d.get("Ocorrência")
        codigos = normalize_occurrence(occ)
        rec["ocorrencia"] = ",".join(codigos) if codigos else 0

        records.append(rec)

    return records


def main():
    if not EXCEL_PATH.exists():
        print(f"ERRO: não encontrei '{EXCEL_PATH.name}' na raiz do projeto.")
        print(f"Coloque a planilha atualizada em: {EXCEL_PATH}")
        sys.exit(1)

    print(f"Lendo {EXCEL_PATH.name}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        print(f"ERRO: a aba '{SHEET_NAME}' não existe nesse arquivo.")
        print(f"Abas disponíveis: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[SHEET_NAME]
    records = clean_records(ws)
    print(f"{len(records)} lançamentos válidos encontrados.")

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write('// Base de dados extraída da planilha Saida_BEL.xlsx, aba "jan_26 (2)"\n')
        f.write(f"// Gerado automaticamente em {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')} — não editar manualmente.\n")
        f.write("const VITLOG_DATA = ")
        json.dump(records, f, ensure_ascii=False)
        f.write(";\n")

    print(f"Pronto! {OUTPUT_PATH.relative_to(PROJECT_ROOT)} atualizado.")
    print("Agora é só atualizar (F5) a página do dashboard no navegador.")


if __name__ == "__main__":
    main()
